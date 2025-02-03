import pandas as pd
from openpyxl import load_workbook
from openpyxl.styles import Font, Alignment
from openpyxl.utils import get_column_letter
from openpyxl.styles import numbers

dtypes = {
    'Consecutivo de registro': 'int64',
    'Número de la resolucion': str,
    'Número de resolucion anterior': str,
    'Fecha de la resolución': str,  
    'Código del tipo de resolución': str,
    'Fecha hasta en suspensiones': str,
    'Número Comparendo': str,
    'Fecha Comparendo': str,  
    'NIP del infractor': str,
    'Código del tipo documento': str,
    'Nombre del infractor': str,
    'Apellido del infractor': str,
    'Direccion del infractor': str,
    'Telefono del Infractor': str,
    'Codigo de la ciudad residencia': str,
    'Valor total de la resolución': str,
    'Valores adicionales.': str,
    'Fotomulta S o N': str,
    'Código organismo que reporta': str,
    'Comparendo Policia de carreteras S o N': str,
    'Código de infracción(*)': str,
    'Valor de la infracción(*)': str,
    'Valor a pagar infraccion(*)': str,
    'Grado de alcoholemia': str,
    'Horas comunitarias': str,
}

ruta_base = "base.xlsx"
ruta_busqueda = "busqueda.xlsx"
ruta_salida = "final.xlsx"
ruta_no_encontrados = "no_encontrados.xlsx"
ruta_duplicados = "duplicados.xlsx"

try:
    df_base = pd.read_excel(ruta_base, dtype=dtypes)
    df_busqueda = pd.read_excel(ruta_busqueda, dtype={'NUMERO_COMPARENDO': str, 'NUMERO_RESOLUCION': str})
    print("Archivos cargados correctamente")
except Exception as e:
    print(f"Error al cargar los archivos: {e}")
    exit()

# Filtrar df_base para incluir solo los comparendos que están en df_busqueda
df_base_filtrada = df_base[df_base['Número Comparendo'].isin(df_busqueda['NUMERO_COMPARENDO'])]

# Identificar duplicados solo en los registros que están en búsqueda
duplicados_base = df_base_filtrada[df_base_filtrada.duplicated(['Número Comparendo'], keep=False)]

if not duplicados_base.empty:
    print(f"Se encontraron {len(duplicados_base)} registros duplicados en la base para los comparendos buscados")
    
    # Ordenar duplicados por número de comparendo para mejor visualización
    duplicados_base = duplicados_base.sort_values('Número Comparendo')
    
    with pd.ExcelWriter(ruta_duplicados, engine='openpyxl') as writer:
        duplicados_base.to_excel(writer, index=False)
        
    wb = load_workbook(ruta_duplicados)
    ws = wb.active
    for column in ws.columns:
        max_length = 0
        column_letter = get_column_letter(column[0].column)
        for cell in column:
            try:
                cell.alignment = Alignment(horizontal='right')
                if cell.row == 1:
                    cell.font = Font(bold=True)
                if len(str(cell.value)) > max_length:
                    max_length = len(str(cell.value))
            except:
                pass
        adjusted_width = (max_length + 2)
        ws.column_dimensions[column_letter].width = adjusted_width
    wb.save(ruta_duplicados)

# Función para seleccionar el registro más reciente
def seleccionar_registro_mas_reciente(grupo):
    try:
        grupo['Fecha Comparendo'] = pd.to_datetime(grupo['Fecha Comparendo'], errors='coerce')
        return grupo.sort_values('Fecha Comparendo', ascending=False).iloc[0]
    except Exception as e:
        print(f"Error al procesar fechas para el comparendo {grupo['Número Comparendo'].iloc[0]}: {e}")
        return grupo.iloc[0]

# Eliminar duplicados solo en los registros filtrados
df_base_sin_duplicados = pd.concat([
    df_base_filtrada.groupby('Número Comparendo').apply(seleccionar_registro_mas_reciente),
    df_base[~df_base['Número Comparendo'].isin(df_busqueda['NUMERO_COMPARENDO'])]
]).reset_index(drop=True)

try:
    resultados = pd.merge(
        df_busqueda[['NUMERO_COMPARENDO', 'NUMERO_RESOLUCION', 'FECHA_RESOLUCION']], 
        df_base_sin_duplicados,
        left_on='NUMERO_COMPARENDO',
        right_on='Número Comparendo',
        how='left'
    )
    
    no_encontrados = df_busqueda[~df_busqueda['NUMERO_COMPARENDO'].isin(df_base['Número Comparendo'])]

    if not no_encontrados.empty:
        with pd.ExcelWriter(ruta_no_encontrados, engine='openpyxl') as writer:
            no_encontrados.to_excel(writer, index=False)
            
        wb = load_workbook(ruta_no_encontrados)
        ws = wb.active
        
        for column in ws.columns:
            max_length = 0
            column_letter = get_column_letter(column[0].column)
            for cell in column:
                try:
                    cell.alignment = Alignment(horizontal='right')
                    if cell.row == 1:
                        cell.font = Font(bold=True)
                    if len(str(cell.value)) > max_length:
                        max_length = len(str(cell.value))
                except:
                    pass
            adjusted_width = (max_length + 2)
            ws.column_dimensions[column_letter].width = adjusted_width
        
        wb.save(ruta_no_encontrados)
        print(f"Se encontraron {len(no_encontrados)} comparendos no encontrados en la base de datos")
        print(f"Los comparendos no encontrados se han guardado en '{ruta_no_encontrados}'")
    
    resultados = resultados.dropna(subset=['Número Comparendo'])
    print("Búsqueda realizada con éxito")
except Exception as e:
    print(f"Error al realizar la búsqueda: {e}")
    exit()

# Duplicar y modificar los resultados
try:
    registros_originales = resultados.copy()
    registros_modificados = resultados.copy()
    
    registros_originales['Código del tipo de resolución'] = '1'
    
    registros_modificados['Número de resolucion anterior'] = registros_modificados['Número de la resolucion']
    registros_modificados['Número de la resolucion'] = registros_modificados['NUMERO_RESOLUCION']
    registros_modificados['Código del tipo de resolución'] = '16'
    registros_modificados['Fecha de la resolución'] = registros_modificados['FECHA_RESOLUCION']
    
    resultados_finales = pd.DataFrame()
    for i in range(len(registros_originales)):
        resultados_finales = pd.concat([
            resultados_finales,
            pd.DataFrame([registros_originales.iloc[i]]),
            pd.DataFrame([registros_modificados.iloc[i]])
        ])
    
    resultados_finales = resultados_finales.drop(['NUMERO_COMPARENDO', 'NUMERO_RESOLUCION', 'FECHA_RESOLUCION'], axis=1)
    print("Registros duplicados y modificados correctamente")
except Exception as e:
    print(f"Error al duplicar y modificar los registros: {e}")
    exit()

try:
    with pd.ExcelWriter(ruta_salida, engine='openpyxl') as writer:
        resultados_finales.to_excel(writer, index=False)

    wb = load_workbook(ruta_salida)
    ws = wb.active
    column_widths = {}

    for row in ws.rows:
        for cell in row:
            if cell.value:
                column_letter = get_column_letter(cell.column)
                current_width = len(str(cell.value))
                if column_letter not in column_widths:
                    column_widths[column_letter] = current_width
                else:
                    if current_width > column_widths[column_letter]:
                        column_widths[column_letter] = current_width

    for row in ws.rows:
        for cell in row:
            cell.alignment = Alignment(horizontal='right')
            if cell.row == 1:
                cell.font = Font(bold=True)
            if isinstance(cell.value, pd.Timestamp):  
                cell.number_format = numbers.FORMAT_DATE_DDMMYYYY

    for column_letter, width in column_widths.items():
        adjusted_width = width + 2
        ws.column_dimensions[column_letter].width = adjusted_width

    wb.save(ruta_salida)

    print(f"Proceso completado. Los resultados se han guardado en '{ruta_salida}'")
    print(f"Total de registros procesados: {len(resultados)}")
    print(f"Total de registros en archivo final: {len(resultados_finales)}")
except Exception as e:
    print(f"Error al guardar el archivo de salida: {e}")